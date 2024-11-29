import pandas as pd
import numpy as np
import openpyxl as opx
from openpyxl.styles import PatternFill



def condition(row):
    '''Условие для столбца "Исчислено всего по формуле".'''
    if row['Налоговая база'] <= 5000000:
        return int(row['Налоговая база'] * 0.13)
    else:
        return int(row['Налоговая база'] * 0.15)

def painting_table(file_name, count_rows):
    '''Покраска шапки и столбца "Исполнения"'''

    # Звгрузка файла
    wb = opx.load_workbook(file_name)
    ws = wb.active

    # Шапка
    for letter in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws[f'{letter}1'].fill = PatternFill(fgColor='3ab3eb',  fill_type='solid')

    # Отклонения
    for i in range(2, count_rows + 2):
        if ws[f'F{i}'].value == 0:
            ws[f'F{i}'].fill = PatternFill(fgColor='79e979',  fill_type='solid') # Зеленый
        else:
            ws[f'F{i}'].fill = PatternFill(fgColor='e76558', fill_type='solid') # Красный

    # Сохранение
    wb.save(file_name)


def build_new_report(data, file_name):
    '''Создание нового, сформулированного отчета от исходного excel файла'''
    df = pd.read_excel(data)

    # Редактирование
    df.drop(labels=['Доход', 'Вычеты', 'Unnamed: 6'], axis=1, inplace=True) # Удаление не нужных столбцов
    df.drop(labels=[0, 1], axis=0, inplace=True) # Удаление первых не нужных строк
    df.dropna(inplace=True)
    df.astype(dtype={'Налог': np.int32})
    df.rename(columns={'Налог': 'Исчислено всего'}, inplace=True)

    # Добавление новых столбцов
    df['Исчислено всего по формуле'] = df.apply(condition, axis=1)
    df['Отклонения'] = df['Исчислено всего'] - df['Исчислено всего по формуле']

    # Сортировка
    df.sort_values(by='Отклонения', ascending=False, inplace=True)

    # Сохранение
    df.to_excel(file_name, index=False)

    # Покраска
    count_rows = df.shape[0]
    painting_table(file_name, count_rows)

